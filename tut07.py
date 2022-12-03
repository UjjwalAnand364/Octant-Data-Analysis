# importing required libraries
import os
os.chdir('C:\\Users\\dell\\Documents\\GitHub\\2001CB60_2022\\tut07')
# path1 = 'C:\\Users\\dell\\Documents\\GitHub\\2001CB60_2022\\tut07\\input'
path2 = 'C:\\Users\\dell\\Documents\\GitHub\\2001CB60_2022\\tut07\\output' # command to create a new folder 'output' in the folder tut06, where all files will be exported.
if os.path.exists(path2)==False:
	os.mkdir(path2)

from datetime import datetime
start_time = datetime.now()

from platform import python_version
ver = python_version()
if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

import multiprocessing 
from multiprocessing import Pool
from itertools import product

import glob
import pandas as pd
import numpy as np
from pandas.io.formats.excel import ExcelFormatter
import openpyxl
from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Color, Font, PatternFill,Border, Side
from openpyxl.styles.differential import DifferentialStyle 
pd.io.formats.excel.ExcelFormatter.header_style = None # to remove the first bold column of the output excel file.
from math import ceil

def octant_analysis(mod,filename):
	reader = pd.read_excel(f'{filename}') # reading the input file and creating a dataframe 'reader'
	size = reader['U'].size # finding no. of elements in one column 
	# returning required values
	df,df_basic,octant_value=octant_range_names(reader,mod,size) 
	df_basic['']=''
	df2,b=octant_longest_subsequence_count_with_range(df_basic,size,octant_value) # b is the row number for longest subsequence length with time ranges.
	df3=overall_transition(df,mod,size)
	df4=mod_transition(df,mod,size)

	goto=ceil(size/mod)# goto is the row length of a particular table.
	df5=df4.iloc[6+goto:,12:21] # df5 is the octant transition counts tables that would be cut and pasted on the right of the dataframe to the right of the rank tables. (original dataframe df4 remains unchanged.) 
	df5=df5.reset_index() # resetting index of df5

	list2=df5.columns.values.tolist() # column names of df5
	list3=['Octant #',1,-1,2,-2,3,-3,4,-4] # column names where df5 is to be inserted
	for i in range(9):
		df4.insert(31+i,list3[i],df5[list2[i+1]]) # inserting at (31+i)th row index and list3[i]th column name.
	for r in range(goto+3,goto+3+16*(goto+1)):
		for c in range(11,21):
			df4.iloc[r,c]='' # removing df4's original octant transition values.
	df4.insert(11,'  ','') # blank columns inserted
	df4.insert(32,'   ','')
	df4.insert(32,'    ','')
	df4.insert(43,'     ','')
	df4.iloc[1,12]=f'Mod {mod}' # inserting input mod value.
	df4.iloc[0,33]='From' # inserting the string "From" at the required cells.
	for i in range(goto): 
		df4.iloc[15+14*i,33]='From' 
	os.chdir(path = 'C:\\Users\\dell\\Documents\\GitHub\\2001CB60_2022\\tut07\\output') # change directory path to point to output folder.
	df4.to_excel(f'{filename[:-5]}_octant_analysis _mod_{mod}.xlsx',index=False) # output the df4 as excel file.

	# In this section, the existing excel file is conditionally formatted, rows inserted and column sizes have been changed as required in the output file.
	wb=openpyxl.load_workbook(f'{filename[:-5]}_octant_analysis _mod_{mod}.xlsx')
	ws=wb.active
	yellowFill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
	ws.insert_rows(1,2) # insert two rows on top of the excel file
	ws['AW1']='Longest Subsequence Length with Range' # inserting headers to each of the different analysis done.
	ws['AS1']='Longest Subsequence Length'
	ws['AI1']='Overal Transition Count'
	ws['AJ2']='To'
	ws['N1']='Overal Octant Count'
	ws.column_dimensions['N'].width=17 # changing column dimensions a/c to the sample output file.
	ws.column_dimensions['AD'].width=27
	ws.column_dimensions['AE'].width=25
	ws.column_dimensions['AF'].width=23
	ws.column_dimensions['AI'].width=22
	ws.column_dimensions['AS'].width=22
	ws.column_dimensions['AT'].width=22
	ws.column_dimensions['AW'].width=28
	ws.column_dimensions['AX'].width=26

	for i in range(2,goto+3):  # setting borders along with coloring the highest rank (i.e. rank 1) using conditional formatting
		ws.conditional_formatting.add(f'W{i+2}:AD{i+2}',CellIsRule(operator='equal', formula=['1'], stopIfTrue=True, fill=yellowFill))

	for i in range(1,10):
		if i>1:
			dxf = DifferentialStyle(fill=yellowFill)
			r = Rule(type="expression", dxf=dxf, formula=[f"=AJ{i+2}=MAX($AJ{i+2}:$AQ{i+2})"],stopIfTrue=True)
			ws.conditional_formatting.add(f'AJ{i+2}:AQ{i+2}', r)
			set_border(ws,f'AI{i+2}:AQ{i+2}') # setting the border along with the coloring the largest element in the overall octant transition count tables.
		else:
			set_border(ws,f'AI{i+2}:AQ{i+2}')
	for a in range(goto): 
		for i in range(16+14*a,25+14*a):
			if i>16+14*a:
				dxf = DifferentialStyle(fill=yellowFill)
				r = Rule(type="expression", dxf=dxf, formula=[f"=AJ{i+2}=MAX($AJ{i+2}:$AQ{i+2})"],stopIfTrue=True)
				ws.conditional_formatting.add(f'AJ{i+2}:AQ{i+2}', r)
				set_border(ws,f'AI{i+2}:AQ{i+2}') # setting the border along with the coloring the largest element in the mod octant transition count tables.
			else:
				set_border(ws,f'AI{i+2}:AQ{i+2}')
	# setting border on each of the other required cells.
	for i in range(3,goto+5):
		set_border(ws,f'N{i}:AF{i}')
	for i in range(3,12):
		set_border(ws,f'AS{i}:AU{i}')
	for i in range(3,b+4):
		set_border(ws,f'AW{i}:AY{i}')
	for i in range(12,21):
		set_border(ws,f'AD{i}:AF{i}')
	
	wb.save(f'{filename[:-5]}_octant_analysis _mod_{mod}.xlsx') # saving the final file

def set_border(ws, cell_range): # function for setting border which takes the worksheet and cell range to set borders on.
	thin = Side(border_style="thin", color="000000") # thickness of border, here set to 'thin'
	for row in ws[cell_range]:
		for cell in row:
			cell.border = Border(top=thin, left=thin, right=thin, bottom=thin) 

def octant_longest_subsequence_count_with_range(df_1,size,octant_value):
	df_1,df_2,store_length2,count,sum=longest_subsequence_length(df_1,octant_value,size) # calling function that adds the columns "Octant Type", "Longest Subsequence Length" and "Count"
	df_1[' ']='' # one column spacing
	df_1['Octant ###']='' # Column headings for time ranges.
	df_1['Longest subsequence length']=''
	df_1['count']=''
	octant_type=[1,-1,2,-2,3,-3,4,-4] # list containing octant types
	a=0 # two variables 'a' and 'k' defined to work on indexes in df_1 and df_2 respectively.
	k=0

	#changing 1--> 1,-1--> 2,2--> 3,-2--> 4 and so on in df_2 and sorting according to this indexing. This is done to improve accessiblity in df2 using minimum code and variables.
	for i in range(sum): # runs until the last column of df_2 is reached
		for j in range(8): # this for-loop finds the octant type.
			if df_2.loc[i,'type']==octant_type[j]:
				df_2.loc[i,'type']=j+1 
				break # breaks inner loop to move on to next row of df_2.
	df_2=df_2.sort_values(by='type',ignore_index=True) # sorting according to newly formed df_2 column 'type'.

	# Main code for time ranges. The approach is to match each row of the column 'type' of df_2, with the variable 'j+1' and then store the time range already present in df2, to df.
	for i in range(sum): 
		for j in range(8): # for-loop run for iterating over each octant type and comparing with the one stored in the column df2['type'] (they are not the usual 1,-1,2,-2 etc. because they have been changed #line 91)
			if df_2.loc[k,'type']==j+1: # if structure- using variable 'k' to compare the values in df_2 with variable 'j+1'
				df_1.loc[a,'Octant ###']=octant_type[j] # storing the octant type, its longest subsequence length and count, as given in sample output file.
				df_1.loc[a,'Longest subsequence length']=store_length2[j]
				df_1.loc[a,'count']=count[j]
				a+=1
				df_1.loc[a,'Octant ###']='Time' # writing strings 'Time', 'From' and 'To' as given in sample output file.
				df_1.loc[a,'Longest subsequence length']="From"
				df_1.loc[a,'count']='To'
				a+=1
				while df_2.loc[k,'type']==j+1: # while loop to store the time ranges of longest subsequence length of a particular octant type( the octant type has been assigned as numbers from 1 to 8 as written in line 91.)
					df_1.loc[a,'Longest subsequence length']=df_2.loc[k,'time1'] # time1 is starting time of longest subsequence length
					df_1.loc[a,'count']=df_2.loc[k,'time2'] # time2 is the ending time of longest subsequence length 
					a+=1
					k+=1
					if k==sum: # if statement to break the while loop once k is equal to sum, because if not done, the loop will repeat for k=sum, and indexing in df2 is from 0 to sum-1.
						break
				break # break to exit out of inner for loop
		if k==sum:
			break
	return df_1,a

def longest_subsequence_length(df_1,octant_value,size):
	octant_type=[1,-1,2,-2,3,-3,4,-4] # list containing octant types
	for i in range(8): # creating the column 'Octant type' corresponding to which longest subsequence length would be stored.
		df_1.loc[i,'Octant ##']=octant_type[i]
	df_1['Longest Subsequence Length']='' # Column headings 
	df_1['Count']=''
	#The lists created below are in one-to-one correspondence with octant_type list.
	length=[1,1,1,1,1,1,1,1] # list to store the repetition of a octant type, which becomes 1 once the repetition is broken 
	store_length=[1,1,1,1,1,1,1,1] # list to store the maximum repetition occured.
	for i in range(size-1): # for loop on the entire column(list) of Octant_value
		for j in range(8): # for loop that runs on the list octant_type to find the type of Octant_value[i] 
			if df_1.loc[i,'Octant_value']==octant_type[j]: # if runs when the correct octant type stored in octant_value[i] is found
				if df_1.loc[i,'Octant_value']==df_1.loc[i+1,'Octant_value']: # if subsequent elements of the octant_value is same, the corresponding octant_type count is incremented
					length[j]+=1
					break # once incremented, the inner for-loop is broken to move on to compare the next two elements.
				 # this condition stores the maximum repetitions of a particular octant_type. Runs only when subsequent repetition is broken or does not occur.
				if store_length[j]<length[j]:
					store_length.pop(j)
					store_length.insert(j,length[j])
				length[j]=1 # after storing the maximum repetitions of a octant type in list store_length, the count of that octant type becomes 1, for the next repetition if it occurs.
				break # breaks the inner for loop to move to compare the next two elements.
	# The following for loop occurs only when a particular octant type is not present even once (which otherwise would be 1 because above we have the entire store_length elements as 1.)
	for j in range(8): 
		if octant_type[j] not in [octant_value[i] for i in range(size-1)]:
			store_length[j]=0
	for j in range(8): # for loop to add the longest subsequent lengths to the dataframe.
		df_1.loc[j,'Longest Subsequence Length']=store_length[j]

	store_length2=store_length # creating a duplicate of store_length 
	# The above code is repeated again, except that now compared to line 88, the if condition also allows equality. This is done to store the number of counts that the
	# longest subsequent length occurs.
	length=[1,1,1,1,1,1,1,1]
	store_length=[1,1,1,1,1,1,1,1]
	count=[0,0,0,0,0,0,0,0] # new list to store how many times the longest subsequent length occurs.

	df_2=pd.DataFrame() #creating new dataframe to store time values of maximum subsequences of each octant type.
	a=0 # new variable to increase row number for storing in df2.
	df_2['type']='' # creating columns in df2
	df_2['time1']='' 
	df_2['time2']=''

	for i in range(size-1): # for loop on the entire column(list) of Octant_value
		for j in range(8): # for loop that runs on the list octant_type to find the type of Octant_value[i] 
			if df_1.loc[i,'Octant_value']==octant_type[j]:
				if df_1.loc[i,'Octant_value']==df_1.loc[i+1,'Octant_value']:
					length[j]+=1
					break
				if store_length[j]<=length[j]:
					store_length.pop(j)
					store_length.insert(j,length[j])
					# when a particular octant type has the longest subsequent length once or more than once, the count list stores it.
					if store_length[j]==store_length2[j]: 
						count[j]+=1
						# the following 3 lines stores the time range in df2 corresponding to the maximum subsequence length for each octant type.
						df_2.loc[a,'type']=octant_type[j]
						df_2.loc[a,'time1']=df_1.loc[i-store_length2[j]+1,'Time']
						df_2.loc[a,'time2']=df_1.loc[i,'Time']
						a+=1 # once above values are stored in df2, a increases and the next row in df2 is used. 
				length[j]=1
				break
	for j in range(8): # for loop to add 'Count' to the dataframe.
		df_1.loc[j,'Count']=count[j]
	sum1=sum(count) # 'sum1' is the no. of rows in df2.
	return df_1,df_2,store_length2,count,sum1 #returns to the function the required arguments.

def overall_transition(df,mod,size):
	# setting index and making a list containing row and column indexes(which are also the octant values).
	i=4+ceil((size/mod)) 
	col_row_names=[1,-1,2,-2,3,-3,4,-4]
	df.loc[i-1,'Octant ID: ']='Overall Transition Count' # strings added as shown in sample output file 
	df.loc[i,'+1']='To'
	df.loc[i+1,'Octant ID: ']='Octant #'
	df.loc[i+2,'']='From'
	# writing row and column headings of the 8x8 matrix
	for j in range(13,21):
		df.iloc[i+1,j]=col_row_names[j-13]
	for j in range(8):
		df.loc[i+2,'Octant ID: ']=col_row_names[j]
		i+=1
	if mod!=size:
		for r in range(10,18): # replacing NaN elements of the 8x8 matrix with integer 0.
			for c in range(13,21):
				df.iloc[r,c]=0 #using iloc to access dataframe cells.
		# code for filling the required transition counts in the 8x8 matrix
		for X in range(size):
			if X==size-1: #if the last element is encountered, entire loop breaks.
				break
			for r in range(10,18): #setting up nested for loop for checking the Xth and (X+1)th element and then incrementing the corresponding cell of the 8x8 matrix
				for c in range(8):
					if df.loc[X,'Octant_value']==col_row_names[r-10] and df.loc[X+1,'Octant_value']==col_row_names[c]: # condition to check the value of X and (X+1)th and element
						df.iloc[r,c+13]+=1 # increments the corresponding cell of 8x8 matrix
	else: # runs only when mod is entered as size, in which case, the usual row indices give incorrect results with above code.
		for r in range(7,15): # replacing NaN elements of the 8x8 matrix with integer 0.
			for c in range(13,21):
				df.iloc[r,c]=0 #using iloc to access dataframe cells.
		# code for filling the required transition counts in the 8x8 matrix
		for X in range(size):
			if X==size-1: #if the last element is encountered, entire loop breaks.
				break
			for r in range(8): #setting up nested for loop for checking the Xth and (X+1)th element and then incrementing the corresponding cell of the 8x8 matrix
				for c in range(8):
					if df.loc[X,'Octant_value']==col_row_names[r] and df.loc[X+1,'Octant_value']==col_row_names[c]: # condition to check the value of X and (X+1)th and element
						df.iloc[r+7,c+13]+=1 # increments the corresponding cell of 8x8 matrix

	return df # returns the resulting dataframe

def mod_transition(df,mod,size):
	col_row_names=[1,-1,2,-2,3,-3,4,-4] # making a list containing the row and column indexes
	# iterating over range of (size//mod)*mod in one iteration; 
	# The variable 'j' serves the purpose of moving down, once every 0->mod-1 elements are covered.
	for j in range(size//mod): 
		i=4+ceil(size/mod)+14*(j+1) # row index for forming the 8x8 matrix and incrementing the row index for each range.
		df.loc[i,'Octant ID: ']='Mod Transition Count' # strings added as shown in sample output file 
		df.loc[i+1,'Octant ID: ']=f'{mod*(j)}->{mod*(j+1)-1}'
		df.loc[i+1,'+1']='To'
		df.loc[i+2,'Octant ID: ']='Octant #'
		df.loc[i+3,'']='From'
		# headings of 8x8 matrix
		for k in range(13,21):
			df.iloc[i+2,k]=col_row_names[k-13]
		for k in range(8):
			df.loc[i+3,'Octant ID: ']=col_row_names[k]
			i+=1
		
		for r in range(11,19): 
			for c in range(13,21):
				df.iloc[r+14*(j+1),c]=0 # replacing NaN elements of the 8x8 matrix with integer 0 using for loop.
		# if else set up for the case when the 'mod' value is a factor of the size of the dataset.
		if size%mod==0: #When the last iteration of the loop is being run and 'mod' is a factor of 'size'.
			if size==mod: # if mod is equal to size
				for r in range(8): 
					for c in range(8):
						df.iloc[r+18,c+13]=0 # replacing NaN elements of the 8x8 matrix with integer 0 using for loop.

				for X in range(mod*j,mod*(j+1)): 
					if X==mod*(j+1)-1:
						break
					for r in range(8): #setting up nested for loop for checking the Xth and (X+1)th element and then incrementing the corresponding cell of the 8x8 matrix
						for c in range(8):
							if df.loc[X,'Octant_value']==col_row_names[r] and df.loc[X+1,'Octant_value']==col_row_names[c]:  # condition to check the value of X and (X+1)th and element
								df.iloc[r+18,c+13]+=1 # increments the corresponding cell of 8x8 matrix
			else:
				for X in range(mod*j,mod*(j+1)): 
					if X==mod*(j+1)-1:
						break
					for r in range(8): #setting up nested for loop for checking the Xth and (X+1)th element and then incrementing the corresponding cell of the 8x8 matrix
						for c in range(8):
							if df.loc[X,'Octant_value']==col_row_names[r] and df.loc[X+1,'Octant_value']==col_row_names[c]:  # condition to check the value of X and (X+1)th and element
								df.iloc[r-3+14*(j+2),c+13]+=1 # increments the corresponding cell of 8x8 matrix

		if size%mod!=0: #When 'mod' is not a factor of 'size'.
			for X in range(mod*j,mod*(j+1)):
				for r in range(8): #setting up nested for loop for checking the Xth and (X+1)th element and then incrementing the corresponding cell of the 8x8 matrix
						for c in range(8):
							if df.loc[X,'Octant_value']==col_row_names[r] and df.loc[X+1,'Octant_value']==col_row_names[c]:  # condition to check the value of X and (X+1)th and element
								df.iloc[r-3+14*(j+2),c+13]+=1 # increments the corresponding cell of 8x8 matrix

	# Forming the 8x8 matrix for the last range (when mod is not a factor of size) and performing the same operations as above.
	if size%mod!=0:
		j+=1
		i=4+ceil(size/mod)+14*(j+1)
		df.loc[i,'Octant ID: ']='Mod Transition Count'
		df.loc[i+1,'Octant ID: ']=f'{mod*(j)}->{size-1}'
		df.loc[i+1,'+1']='To'
		df.loc[i+2,'Octant ID: ']='Octant #'
		df.loc[i+3,'']='From'
		for k in range(13,21):
			df.iloc[i+2,k]=col_row_names[k-13]
		for k in range(8):
			df.loc[i+3,'Octant ID: ']=col_row_names[k]
			i+=1

		for r in range(11,19):
			for c in range(13,21):
				df.iloc[r+14*(j+1),c]=0

		for X in range(mod*j,size):
			if X==size-1:
				break
			for r in range(8):
				for c in range(8):
					if df.loc[X,'Octant_value']==col_row_names[r] and df.loc[X+1,'Octant_value']==col_row_names[c]:
						df.iloc[r-3+14*(j+2),c+13]+=1
	
	return df

def octant_range_names(reader,mod,size):
	size = reader['U'].size # finding no. of elements in one column 
	u_avg = reader['U'].mean() # mean of 'U' column
	v_avg = reader['V'].mean() # mean of 'V' column
	w_avg = reader['W'].mean() # mean of 'W' column

	df = pd.DataFrame() # creating a dataframe object that stores each of the lists created. 
	df.loc[:,"Time"]=reader['T']
	df.loc[:,'U']=reader["U"]
	df.loc[:,'V']=reader["V"]
	df.loc[:,'W']=reader["W"]
	df.loc[0,'U avg']=u_avg
	df.loc[0,'V avg']=v_avg
	df.loc[0,'W avg']=w_avg

	# lists to store the U-u_avg, V-v_avg, W-w_avg values
	list1 = []
	list2 = []
	list3 = []
	octant_value = [] # list to store the octant values

	# appending U-u_avg, V-v_avg, W-w_avg values in the respective lists
	for i in range(size): # iterating over each element in 'U' column and subtracting the u_avg from them, and then storing in list1.
		list1.append(reader['U'][i]-u_avg)
	for i in range(size): # iterating over each element in 'V' column and subtracting the v_avg from them, and then storing in list2.
		list2.append(reader['V'][i]-v_avg)
	for i in range(size): # iterating over each element in 'W' column and subtracting the w_avg from them, and then storing in list3.
		list3.append(reader['W'][i]-w_avg)

	df.loc[:,"U - U avg"]=list1
	df.loc[:,"V - V avg"]=list2
	df.loc[:,"W - W avg"]=list3

	# Now we have completed data pre-processing
	
	# assigning octant value to each U - U avg, V - V avg , W - W avg row using if-else structure and the octant rule.
	for i in range(size): 
		if df['W - W avg'][i] < 0: # if  is less than 0, octant is negative
			if df['V - V avg'][i] >= 0: 
				if df['U - U avg'][i] >= 0:
					octant_value.append(-1) 
				else:
					octant_value.append(-2)
			else:
				if df['U - U avg'][i] >= 0:
					octant_value.append(-4)
				else:
					octant_value.append(-3)
		else:
			if df['V - V avg'][i] >= 0:
				if df['U - U avg'][i] >= 0:
					octant_value.append(1)
				else:
					octant_value.append(2)
			else:
				if df['U - U avg'][i] >= 0:
					octant_value.append(4)
				else:
					octant_value.append(3)
	
	df["Octant_value"] = octant_value    # storing the octant values in a new column named 'Octant_value'
	df_basic=df
	df.loc[1,'']="User input: " # creating one column to display the string 'user input'
	count0 = [0, 0, 0, 0, 0, 0, 0, 0] # creating a list to store the count of each type of octant. 8 zeroes correspond to 8 different octant IDs.

	# using for loop on the 'Octant_value' column and if-else structure to store the number of each type of octant ID. 
	for i in range(size):
		if df['Octant_value'][i] == 1:
			count0[0] += 1
		elif df['Octant_value'][i] == -1:
			count0[1] += 1
		elif df['Octant_value'][i] == 2:
			count0[2] += 1
		elif df['Octant_value'][i] == -2:
			count0[3] += 1
		elif df['Octant_value'][i] == 3:
			count0[4] += 1
		elif df['Octant_value'][i] == -3:
			count0[5] += 1
		elif df['Octant_value'][i] == 4:
			count0[6] += 1
		elif df['Octant_value'][i] == -4:
			count0[7] += 1
	df.loc[0, 'Octant ID: '] = 'Overall count: ' # creating a new column 'Octant ID' to display the string "overall count".
	# storing the number of each type of octant ID using the values stored in the list.
	df.loc[0, "+1"] = count0[0]
	df.loc[0, "-1"] = count0[1]
	df.loc[0, "+2"] = count0[2]
	df.loc[0, "-2"] = count0[3]
	df.loc[0, "+3"] = count0[4]
	df.loc[0, "-3"] = count0[5]
	df.loc[0, "+4"] = count0[6]
	df.loc[0, "-4"] = count0[7]
	# creating a separate function with a recursion to do Task 2 as specified in the problem.
	count2 = 0 # this variable iterates from 0 to mod value, and then gets increased by mod in the next recursion.
	i = 1 # this is used to add elements at column at ith index
	count3 = [0, 0, 0, 0, 0, 0, 0, 0] # counts each type of octant ID
	df = mod_count(df, i, count2, size, mod, count3) # calling the function which takes 6 arguments.

	octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
	octant_type=[1,-1,2,-2,3,-3,4,-4] # list of octant types
	goto=1+ceil(size/mod) # this goto variable is the maximum row index upto which we have to perform octant rank operation
	for i in range(8):
		df[f'Rank of {octant_type[i]}']='' # creating 8 columns with heading: Rank of 1, Rank of 2 etc.
	df['Rank1 Octant ID']='' # 2 extra columns with given names
	df['Rank1 Octant Name']=''

	for j in range(0,goto): # for-loop to continue till the maximum row index
		store=df.iloc[j,13:21] # making two identical lists to store the j'th row counts
		store_list=list(store) 
		store_list2=list(store)
		store_list.sort(reverse=True)
		# nested for loops to input the rank of each single count element in the dataframe (as required by the task)
		for i in range(8): # i is used to iterate over the unsorted list which contains elements in the same order as given by count function.
			for k in range(8): # k is used to iterate over the sorted list
				if store_list2[i]==store_list[k]: # the elements of unsorted list are compared with those in the sorted list.
					df.iloc[j,21+i]=k+1 # if ith element of unsorted list matches with kth element of the sorted list, the corresponding position in the sorted list is incremented by 1 and added to the dataframe.
					break # breaks inner for-loop to go to the next element in the unsorted list.
		
		for i in range(8): # this for loop writes the maximum element in the Rank1 octant ID and rank1 octant name.
			if store_list[0]==store_list2[i]: # if the first element of sorted list,matches with the ith element of the unsorted list, the ith element of list octant type and octant_name_id_mapping dictionary will be written. 
				df.iloc[j,29]=octant_type[i]
				df.iloc[j,30]=octant_name_id_mapping[f'{octant_type[i]}']
	# creates the bottom table as required in the output file.
	df.iloc[goto+3,28]='Octant ID' # column names
	df.iloc[goto+3,29]='Octant Name' 
	df.iloc[goto+3,30]='Count of Rank 1 Mod Values'
	for i in range(8): # for loop to simply store the octant type and corresponding octant name.
		df.iloc[goto+4+i,28]=octant_type[i]
		df.iloc[goto+4+i,29]=octant_name_id_mapping[f'{octant_type[i]}']

	count2=[0,0,0,0,0,0,0,0] # list to store the maximum rank of each octant type in mod-counts.
	for i in range(8): # for loop to operate for each row containing mod-count of rank table.
		rank_mod=df.iloc[1:goto,22+i] #store elements present in those mod-counts.
		rank_mod_list=list(rank_mod) #make it a list
		count2[i]=rank_mod_list.count(1) # count the no. of times a particular octant type in each row.
		
	for i in range(8):
		df.iloc[goto+4+i,30]=count2[i] # write the count list to the bottom table
	
	return df,df_basic,octant_value #returns the df dataframe to the main code
	
def mod_count(df, i, count2, size, mod, count3): 
	count3 = [0, 0, 0, 0, 0, 0, 0, 0] # creating a separate list which stores each octant ID value.
	# if-else structure used.
	if count2+mod<size: # if count2+mod<size, we use this code, and when count2+mod is greater than size, use base condition:
		for j in range(count2, count2+mod): # iterates for elements of octant_value ranging from (count2 to count2+mod)
			if df['Octant_value'][j] == 1: 
				count3[0] += 1
			elif df['Octant_value'][j] == -1:
				count3[1] += 1
			elif df['Octant_value'][j] == 2:
				count3[2] += 1
			elif df['Octant_value'][j] == -2:
				count3[3] += 1
			elif df['Octant_value'][j] == 3:
				count3[4] += 1
			elif df['Octant_value'][j] == -3:
				count3[5] += 1
			elif df['Octant_value'][j] == 4:
				count3[6] += 1
			elif df['Octant_value'][j] == -4:
				count3[7] += 1

		df.loc[i, 'Octant ID: '] = f'{count2}->{count2+mod-1}:' # writes the range for which the octant ID values are found
		# displays the count of each octant Id in the range count2,count2+mod
		for k in range(13,21):
			df.iloc[i,k] = count3[k-13]
		count2 += mod # increases the value of count2 so that the octant ID range changes
		i += 1 #increases the value for recursion
		mod_count(df, i, count2, size, mod, count3) #recursion call
		return df #returns the final dataframe to the function octant_identification
	else: # final base condition 
		for j in range(count2, size): # iterates for elements of octant_value for the final range until the last element.
			if df['Octant_value'][j] == 1: 
				count3[0] += 1
			elif df['Octant_value'][j] == -1:
				count3[1] += 1
			elif df['Octant_value'][j] == 2:
				count3[2] += 1
			elif df['Octant_value'][j] == -2:
				count3[3] += 1
			elif df['Octant_value'][j] == 3:
				count3[4] += 1
			elif df['Octant_value'][j] == -3:
				count3[5] += 1
			elif df['Octant_value'][j] == 4:
				count3[6] += 1
			elif df['Octant_value'][j] == -4:
				count3[7] += 1

		df.loc[i, 'Octant ID: '] = f'{count2}->{size-1}:' # prints the final octant ID range
		# displays the count of each octant Id in the range (count2,size)
		for k in range(13,21):
			df.iloc[i,k] = count3[k-13]

	return df #returns to the previous count() call.

try:
	mod=5000
	os.chdir('C:\\Users\\dell\\Documents\\GitHub\\2001CB60_2022\\tut07\\input')
	files=glob.glob('*.xlsx') # getting .xlsx files from the 'input' folder
	for file in files:
		octant_analysis(mod,file)
		os.chdir('C:\\Users\\dell\\Documents\\GitHub\\2001CB60_2022\\tut07\\input') # changing directory to 'input' as inside the function the directory gets changed to 'output'.
except KeyboardInterrupt:
	print('Keyboard Interrupt occured!')
except:
	print("Please verify that the input folder with files exists!")

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
